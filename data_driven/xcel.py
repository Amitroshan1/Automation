import openpyxl

# Replace this with the path to your Excel file
path = "C:\\Users\\dell\\Desktop\\testcases.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.ge

row = sheet.max_row
col = sheet.max_column

print(row)
print(col)

# Example of reading cell values
for row_num in range(1, row+1):
    for col_num in range(1, col+1):
        cell_value = sheet.cell(row=row_num, column=col_num).value
        print(f"Cell ({row_num}, {col_num}): {cell_value}")

workbook.close()
