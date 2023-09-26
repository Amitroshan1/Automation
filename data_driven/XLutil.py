import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(path,sheetname):
    workbook= openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def getColCount(path,sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)

def readData(path,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum,column=colnum).value

def writeData(path,sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum,column=colnum).value=data
    workbook.save(path)

def fillcolor(path,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    greenfill=PatternFill(start_color="#4C8E0E",end_color="#4C8E0E",fill_type='solid')
    sheet.cell(row=rownum,column=colnum).fill=greenfill
    workbook.save(path)

def fillredcolor(path,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    Redfill=PatternFill(start_color='#ef1818',end_color='#ef1818',fill_type='solid')
    sheet.cell(row=rownum,column=colnum).fill=Redfill
    workbook.save(path)